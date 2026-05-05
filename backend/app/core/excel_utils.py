from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel_sheet(worksheet, df, title="BÁO CÁO HỆ THỐNG CRM"):
    if df.empty or len(df.columns) < 1:
        return
    
    # VNPost Colors
    header_fill = PatternFill(start_color='0054A6', end_color='0054A6', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Add Title row
    worksheet.insert_rows(1)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title.upper()
    title_cell.font = Font(color='F9A51A', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    worksheet.row_dimensions[1].height = 30

    # Header styling
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    worksheet.row_dimensions[2].height = 25

    # Column name mapping for optimization
    center_cols = {"STT", "Mã CRM/CMS", "Mã KH", "Mã BC", "Mã Đơn vị", "Loại", "Trạng thái Vòng đời", "Phân khúc RFM", "Hạng RFM", "Nhóm KH", "Trạng thái", "Điểm Sức khỏe (0-100)", "Số điện thoại", "Ngày kết thúc HĐ", "Tần suất gửi (Ngày)", "Ngày giao dịch gần nhất"}
    col_names = list(df.columns)
    
    # Pre-calculate widths
    widths = [len(str(name)) for name in col_names]

    # Data styling (Batch processing rows for speed)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    for row_num, row in enumerate(worksheet.iter_rows(min_row=3, max_row=worksheet.max_row), 3):
        is_alt = (row_num % 2 == 0)
        for i, cell in enumerate(row):
            cell.border = border
            if is_alt:
                cell.fill = alt_fill
            
            col_name = col_names[i]
            if col_name in center_cols:
                cell.alignment = Alignment(horizontal='center')
            
            # Currency formatting
            if any(key in col_name for key in ["Doanh thu", "VNĐ", "Biến động", "diff", "rev_a", "rev_b"]):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

            # Width tracking
            val_str = str(cell.value) if cell.value is not None else ""
            if len(val_str) > widths[i]:
                widths[i] = len(val_str)

    # Apply widths
    for i, w in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = min(w + 2, 50)

    # Freeze panes
    worksheet.freeze_panes = 'A3'
